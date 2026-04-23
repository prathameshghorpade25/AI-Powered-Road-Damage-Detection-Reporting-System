from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

_TRAFFIC_RANK = {"low": 1, "medium": 2, "high": 3, "very_high": 4}

_ZONE_CENTER: dict[str, tuple[float, float]] = {
    "Kopargaon Core": (19.8824, 74.4760),
    "Shirdi Pilgrim Corridor": (19.7645, 74.4774),
    "Rahata Corridor": (19.7155, 74.4800),
    "Ahilyanagar Urban": (19.0952, 74.7496),
    "Sangamner Connector": (19.5670, 74.2090),
    "Regional Arterials": (19.8200, 74.5200),
}

_ZONE_KEYWORDS: list[tuple[str, tuple[str, ...]]] = [
    ("Kopargaon Core", ("kopargaon", "kolpewadi", "godavari bridge", "shivaji chowk", "bus depot")),
    ("Shirdi Pilgrim Corridor", ("shirdi", "airport road", "temple")),
    ("Rahata Corridor", ("rahata", "rahuri", "puntamba")),
    ("Ahilyanagar Urban", ("ahilyanagar", "savedi", "nagar", "midc", "pipeline road")),
    ("Sangamner Connector", ("sangamner", "sinnar", "chitali", "jejuri", "belapur", "pathardi")),
]


def _normalize(value: str) -> str:
    txt = "".join(ch.lower() if (ch.isalnum() or ch.isspace()) else " " for ch in value)
    return " ".join(txt.split())


def _normalize_traffic(value: str) -> str:
    norm = _normalize(value).replace(" ", "_")
    if norm in _TRAFFIC_RANK:
        return norm
    return "medium"


def _zone_for_location(location: str) -> str:
    loc = _normalize(location)
    for zone, words in _ZONE_KEYWORDS:
        if any(word in loc for word in words):
            return zone
    return "Regional Arterials"


def _load_rows() -> list[dict[str, str]]:
    csv_path = Path(__file__).resolve().parents[2] / "authority" / "traffic_data.csv"
    if not csv_path.exists():
        return []
    rows: list[dict[str, str]] = []
    with csv_path.open("r", encoding="utf-8") as f:
        raw = csv.reader(f)
        try:
            header = next(raw)
        except StopIteration:
            return []
        normalized = [_normalize(h).replace(" ", "_") for h in header]
        idx_location = normalized.index("location") if "location" in normalized else 0
        idx_traffic = normalized.index("traffic") if "traffic" in normalized else 1
        idx_time = normalized.index("time") if "time" in normalized else 2
        for row in raw:
            if not row:
                continue
            location = (row[idx_location] if idx_location < len(row) else "").strip()
            traffic = (row[idx_traffic] if idx_traffic < len(row) else "").strip()
            peak_time = (row[idx_time] if idx_time < len(row) else "").strip()
            if not location:
                continue
            rows.append({"location": location, "traffic": traffic, "time": peak_time})
    return rows


def _parse_coordinate(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)", text)
    if not match:
        return None
    val = float(match.group(1))
    if "S" in text or "W" in text:
        val = -abs(val)
    return val


def _load_coordinate_rows() -> list[dict[str, Any]]:
    csv_path = Path(__file__).resolve().parents[2] / "authority" / "addresses_with_coordinates.csv"
    rows_out: list[dict[str, Any]] = []
    if csv_path.exists():
        with csv_path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                location = (row.get("Location") or "").strip()
                if not location:
                    continue
                lat = _parse_coordinate(row.get("Latitude"))
                lon = _parse_coordinate(row.get("Longitude"))
                if lat is None or lon is None:
                    continue
                rows_out.append({"location": location, "location_norm": _normalize(location), "lat": lat, "lon": lon})
        if rows_out:
            return rows_out

    xlsx_path = Path(__file__).resolve().parents[2] / "authority" / "Addresses with coordinates.xlsx"
    if not xlsx_path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_normalize(str(h or "")).replace(" ", "_") for h in rows[0]]
    idx_location = header.index("location") if "location" in header else 1
    idx_lat = header.index("latitude") if "latitude" in header else 2
    idx_lon = header.index("longitude") if "longitude" in header else 3

    for row in rows[1:]:
        location = (str(row[idx_location]).strip() if idx_location < len(row) and row[idx_location] else "")
        if not location:
            continue
        lat = _parse_coordinate(row[idx_lat] if idx_lat < len(row) else None)
        lon = _parse_coordinate(row[idx_lon] if idx_lon < len(row) else None)
        if lat is None or lon is None:
            continue
        rows_out.append({"location": location, "location_norm": _normalize(location), "lat": lat, "lon": lon})
    return rows_out


def _best_coordinate_for_location(location: str, coord_rows: list[dict[str, Any]]) -> tuple[float, float] | None:
    target = _normalize(location)
    if not target:
        return None

    # Exact/containment match first.
    for row in coord_rows:
        cand = str(row["location_norm"])
        if target == cand or target in cand or cand in target:
            return float(row["lat"]), float(row["lon"])

    target_tokens = set(target.split())
    best_score = 0.0
    best: tuple[float, float] | None = None
    for row in coord_rows:
        cand_tokens = set(str(row["location_norm"]).split())
        if not cand_tokens:
            continue
        overlap = len(target_tokens.intersection(cand_tokens))
        score = overlap / len(target_tokens.union(cand_tokens)) if target_tokens else 0.0
        if score > best_score:
            best_score = score
            best = (float(row["lat"]), float(row["lon"]))
    return best if best_score >= 0.2 else None


def build_traffic_zones() -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    coord_rows = _load_coordinate_rows()
    for row in _load_rows():
        grouped[_zone_for_location(row["location"])].append(row)

    zones: list[dict[str, Any]] = []
    for zone_name, rows in grouped.items():
        center = _ZONE_CENTER.get(zone_name, _ZONE_CENTER["Regional Arterials"])
        coords = [_best_coordinate_for_location(r["location"], coord_rows) for r in rows]
        valid_coords = [c for c in coords if c is not None]
        if valid_coords:
            avg_lat = sum(c[0] for c in valid_coords) / len(valid_coords)
            avg_lon = sum(c[1] for c in valid_coords) / len(valid_coords)
            center = (avg_lat, avg_lon)
        ranked = sorted((_normalize_traffic(r["traffic"]) for r in rows), key=lambda x: _TRAFFIC_RANK[x], reverse=True)
        top_level = ranked[0] if ranked else "medium"
        peak_rows = [r for r in rows if _normalize_traffic(r["traffic"]) in ("high", "very_high")]
        peak_time = peak_rows[0]["time"] if peak_rows else (rows[0]["time"] if rows else "")
        radius_m = 450 + len(rows) * 55
        zones.append(
            {
                "name": zone_name,
                "centerLat": center[0],
                "centerLon": center[1],
                "radiusM": radius_m,
                "trafficLevel": top_level,
                "peakTime": peak_time,
                "hotspotCount": len(rows),
                "sampleLocations": [r["location"] for r in rows[:5]],
            }
        )

    zones.sort(key=lambda z: (_TRAFFIC_RANK.get(str(z["trafficLevel"]), 0), int(z["hotspotCount"])), reverse=True)
    return zones


def build_traffic_hotspots() -> list[dict[str, Any]]:
    coord_rows = _load_coordinate_rows()
    grouped: dict[str, dict[str, Any]] = {}
    for row in _load_rows():
        address = row["location"].strip()
        if not address:
            continue
        level = _normalize_traffic(row["traffic"])
        zone = _zone_for_location(address)
        rec = grouped.get(address)
        if rec is None:
            rec = {
                "address": address,
                "zone": zone,
                "trafficLevel": level,
                "peakTime": row["time"],
                "occurrences": 0,
            }
            grouped[address] = rec
        rec["occurrences"] += 1
        if _TRAFFIC_RANK[level] > _TRAFFIC_RANK[str(rec["trafficLevel"])]:
            rec["trafficLevel"] = level
            rec["peakTime"] = row["time"]

    hotspots: list[dict[str, Any]] = []
    used_locations: set[str] = set()
    for address, rec in grouped.items():
        zone = str(rec["zone"])
        coord = _best_coordinate_for_location(str(rec["address"]), coord_rows)
        if not coord:
            continue
        used_locations.add(_normalize(address))
        lat, lon = coord
        occurrences = int(rec["occurrences"])
        hotspots.append(
            {
                "address": rec["address"],
                "zone": zone,
                "lat": round(lat, 6),
                "lon": round(lon, 6),
                "trafficLevel": rec["trafficLevel"],
                "peakTime": rec["peakTime"],
                "occurrences": occurrences,
                "radiusM": 45 + min(occurrences * 8, 40),
            }
        )

    # Include coordinate-sheet locations not present in traffic_data.csv.
    for row in coord_rows:
        norm_loc = str(row["location_norm"])
        if norm_loc in used_locations:
            continue
        zone = _zone_for_location(str(row["location"]))
        hotspots.append(
            {
                "address": row["location"],
                "zone": zone,
                "lat": round(float(row["lat"]), 6),
                "lon": round(float(row["lon"]), 6),
                "trafficLevel": "medium",
                "peakTime": "",
                "occurrences": 1,
                "radiusM": 45,
            }
        )

    hotspots.sort(
        key=lambda h: (_TRAFFIC_RANK.get(str(h["trafficLevel"]), 0), int(h["occurrences"])),
        reverse=True,
    )
    return hotspots

